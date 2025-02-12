#!/usr/bin/env python3
"""
Load & Timesheet Manager

Connects to a PostgreSQL database using SQL settings provided via a startup dialog.
Then it lets you view collection loads (jobs where dwjtype='C') for a selected week 
(week-ending on Sunday) and generate loadsheets (Excel only) and timesheets (Excel only).
You can later convert these Excel files to PDF on demand.
A Mapping tab is provided for setting cell mappings.
SQL, Mapping, and Signature settings are persisted in config.ini.
Window geometry is saved on exit.
All errors are logged to "errorloadsheet.log".
"""

import os
import sys
import subprocess
import logging
import configparser
from datetime import datetime, timedelta

import psycopg2
from psycopg2.extras import RealDictCursor
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook

# -----------------------
# Logging Setup
LOG_FORMAT = "%(asctime)s %(levelname)s: %(message)s"
logging.basicConfig(filename="errorloadsheet.log", level=logging.DEBUG, format=LOG_FORMAT)
logger = logging.getLogger(__name__)

# -----------------------
# Global Paths and Template Names
BASE_DIR = os.getcwd()
LOADSHEET_FOLDER = os.path.join(BASE_DIR, "loadsheets")
TIMESHEET_FOLDER = os.path.join(BASE_DIR, "timesheets")
TEMPLATE_LOADSHEET = os.path.join(BASE_DIR, "LoadsheetTemplate.xlsx")
TEMPLATE_TIMESHEET = os.path.join(BASE_DIR, "TimesheetTemplate.xlsx")
for folder in [LOADSHEET_FOLDER, TIMESHEET_FOLDER]:
    os.makedirs(folder, exist_ok=True)
CONFIG_FILE = os.path.join(BASE_DIR, "config.ini")

# -----------------------
# Utility Functions
def format_date(datestr):
    """Convert a YYYYMMDD string to 'Day / dd/mm/yyyy'."""
    try:
        dt = datetime.strptime(datestr, "%Y%m%d")
        return dt.strftime("%A / %d/%m/%Y")
    except Exception as e:
        logger.exception("Date formatting error for %s", datestr)
        return datestr

def upcoming_sunday():
    """Return a date object for the upcoming Sunday (or today if Sunday)."""
    today = datetime.today()
    days_ahead = 6 - today.weekday()  # Monday=0, Sunday=6
    if days_ahead < 0:
        days_ahead += 7
    return today + timedelta(days=days_ahead)

def convert_excel_to_pdf(excel_path, output_dir):
    """Convert an Excel file to PDF using LibreOffice in headless mode."""
    libreoffice_cmd = "soffice"  # Assumes LibreOffice is in your PATH.
    excel_path = os.path.abspath(excel_path)
    output_dir = os.path.abspath(output_dir)
    cmd = f'"{libreoffice_cmd}" --headless --convert-to pdf:calc_pdf_Export --outdir "{output_dir}" "{excel_path}"'
    logger.debug(f"PDF conversion command: {cmd}")
    try:
        result = subprocess.run(cmd, shell=True, text=True, capture_output=True, timeout=60)
        if result.returncode == 0:
            pdf_path = os.path.join(output_dir, os.path.splitext(os.path.basename(excel_path))[0] + ".pdf")
            if os.path.exists(pdf_path):
                logger.info(f"PDF created: {pdf_path}")
                return pdf_path
            else:
                logger.error("PDF conversion completed but file not found")
        else:
            logger.error(f"PDF conversion error: {result.stderr}")
    except Exception as e:
        logger.error(f"Exception during PDF conversion: {e}")
    return None

def get_unique_filename(folder, prefix, extension):
    """Generate a unique filename using the current timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.{extension}"
    full_path = os.path.join(folder, filename)
    counter = 1
    while os.path.exists(full_path):
        filename = f"{prefix}_{timestamp}({counter}).{extension}"
        full_path = os.path.join(folder, filename)
        counter += 1
    return full_path

# -----------------------
# PostgreSQL Database Manager
class DatabaseManager:
    def __init__(self, host, port, dbname, user, password):
        self.host = host
        self.port = port
        self.dbname = dbname
        self.user = user
        self.password = password
        self.conn = None

    def connect(self):
        try:
            self.conn = psycopg2.connect(
                host=self.host,
                port=self.port,
                dbname=self.dbname,
                user=self.user,
                password=self.password,
                cursor_factory=RealDictCursor
            )
            logger.info("Connected to PostgreSQL successfully.")
            return True
        except Exception as e:
            logger.exception("PostgreSQL connection failed")
            return False

    def fetch_loads(self, start_date, end_date):
        try:
            with self.conn.cursor() as cur:
                query = """
                    SELECT * FROM jobs
                    WHERE dwjdate BETWEEN %s AND %s
                      AND dwjtype = 'C'
                    ORDER BY dwjdate DESC
                """
                cur.execute(query, (start_date, end_date))
                return cur.fetchall()
        except Exception as e:
            logger.exception("Failed to fetch loads")
            return []

    def fetch_destination(self, load_id):
        try:
            with self.conn.cursor() as cur:
                query = "SELECT * FROM jobs WHERE dwjload = %s AND dwjtype = 'D' LIMIT 1"
                cur.execute(query, (load_id,))
                return cur.fetchone()
        except Exception as e:
            logger.exception("Failed to fetch destination for load %s", load_id)
            return None

    def fetch_vehicle_details(self, load_id):
        try:
            with self.conn.cursor() as cur:
                query = """
                    SELECT v.dwvvehref, v.dwvmoddes, c.offloaded, c.docs, c.sparekeys, c.photos, v.dwvkey
                    FROM vehicles v
                    LEFT JOIN car_info c ON v.dwvkey = c.dwvkey
                    WHERE v.dwvload = %s
                """
                cur.execute(query, (load_id,))
                return cur.fetchall()
        except Exception as e:
            logger.exception("Failed to fetch vehicle details for load %s", load_id)
            return []

    def update_vehicle_details(self, dwvkey, new_model, offloaded, docs, sparekeys, photos):
        try:
            with self.conn.cursor() as cur:
                cur.execute("UPDATE vehicles SET dwvmoddes = %s WHERE dwvkey = %s", (new_model, dwvkey))
                cur.execute("""
                    UPDATE car_info
                    SET offloaded = %s, docs = %s, sparekeys = %s, photos = %s
                    WHERE dwvkey = %s
                """, (offloaded, docs, sparekeys, photos, dwvkey))
                self.conn.commit()
        except Exception as e:
            logger.exception("Failed to update vehicle details for %s", dwvkey)
            self.conn.rollback()

# -----------------------
# Configuration Handling
def load_config():
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
    else:
        config["SQL"] = {
            "host": "localhost",
            "port": "5432",
            "dbname": "your_database",
            "user": "postgres",
            "password": "",
            "poll_interval": "60"
        }
        config["Window"] = {"geometry": ""}
        config["SIGNATURE1"] = {"scale": "100", "x_offset": "0", "y_offset": "0"}
        config["SIGNATURE2"] = {"scale": "100", "x_offset": "0", "y_offset": "0"}
        config["LOADSHEET_MAPPING"] = {"date_cell": "C6", "load_cell": "G6", "collection_cell": "B9", "destination_cell": "F9"}
        with open(CONFIG_FILE, "w") as f:
            config.write(f)
    return config

def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        config.write(f)

# -----------------------
# Settings Dialog (SQL Connection)
class SettingsDialog(tk.Toplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("Database Settings")
        self.db_manager = None
        self.config = load_config()
        self.initUI()

    def initUI(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)
        # Use pg_* keys if available; otherwise, fallback.
        host = self.config["SQL"].get("pg_host", self.config["SQL"].get("host", "localhost"))
        port = self.config["SQL"].get("pg_port", self.config["SQL"].get("port", "5432"))
        dbname = self.config["SQL"].get("pg_database", self.config["SQL"].get("dbname", "your_database"))
        user = self.config["SQL"].get("pg_username", self.config["SQL"].get("user", "postgres"))
        password = self.config["SQL"].get("pg_password", self.config["SQL"].get("password", ""))
        ttk.Label(frame, text="Host:", font=("Arial", 14)).grid(row=0, column=0, sticky="e", pady=5)
        self.host_entry = ttk.Entry(frame, font=("Arial", 14))
        self.host_entry.insert(0, host)
        self.host_entry.grid(row=0, column=1, pady=5)
        ttk.Label(frame, text="Port:", font=("Arial", 14)).grid(row=1, column=0, sticky="e", pady=5)
        self.port_entry = ttk.Entry(frame, font=("Arial", 14))
        self.port_entry.insert(0, port)
        self.port_entry.grid(row=1, column=1, pady=5)
        ttk.Label(frame, text="Database:", font=("Arial", 14)).grid(row=2, column=0, sticky="e", pady=5)
        self.dbname_entry = ttk.Entry(frame, font=("Arial", 14))
        self.dbname_entry.insert(0, dbname)
        self.dbname_entry.grid(row=2, column=1, pady=5)
        ttk.Label(frame, text="Username:", font=("Arial", 14)).grid(row=3, column=0, sticky="e", pady=5)
        self.username_entry = ttk.Entry(frame, font=("Arial", 14))
        self.username_entry.insert(0, user)
        self.username_entry.grid(row=3, column=1, pady=5)
        ttk.Label(frame, text="Password:", font=("Arial", 14)).grid(row=4, column=0, sticky="e", pady=5)
        self.password_entry = ttk.Entry(frame, font=("Arial", 14), show="*")
        self.password_entry.insert(0, password)
        self.password_entry.grid(row=4, column=1, pady=5)
        btn = tk.Button(frame, text="Connect", command=self.try_connect, font=("Arial", 14), width=20)
        btn.grid(row=5, column=0, columnspan=2, pady=10)

    def try_connect(self):
        host = self.host_entry.get().strip()
        port = self.port_entry.get().strip()
        dbname = self.dbname_entry.get().strip()
        user = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        self.db_manager = DatabaseManager(host, port, dbname, user, password)
        if self.db_manager.connect():
            messagebox.showinfo("Success", "Connected successfully!")
            # Save both sets of keys for compatibility.
            self.config["SQL"]["pg_host"] = host
            self.config["SQL"]["host"] = host
            self.config["SQL"]["pg_port"] = port
            self.config["SQL"]["port"] = port
            self.config["SQL"]["pg_database"] = dbname
            self.config["SQL"]["dbname"] = dbname
            self.config["SQL"]["pg_username"] = user
            self.config["SQL"]["user"] = user
            self.config["SQL"]["pg_password"] = password
            self.config["SQL"]["password"] = password
            save_config(self.config)
            self.destroy()
        else:
            messagebox.showerror("Error", "Failed to connect to database. Check errorloadsheet.log.")

    def get_db_manager(self):
        return self.db_manager

# -----------------------
# Settings Tab (for updating SQL settings and opening Signature Settings)
class SettingsTab(tk.Frame):
    def __init__(self, parent, config):
        super().__init__(parent)
        self.config = config
        self.create_ui()

    def create_ui(self):
        frame = tk.Frame(self, padx=20, pady=20)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text="SQL Settings", font=("Arial", 16, "bold")).pack(pady=10)
        fields = [("Host", "host"), ("Port", "port"), ("Database", "dbname"),
                  ("Username", "user"), ("Password", "password"), ("Polling Interval", "poll_interval")]
        self.entries = {}
        for label_text, key in fields:
            row = tk.Frame(frame)
            row.pack(fill="x", pady=5)
            tk.Label(row, text=f"{label_text}:", width=15, anchor="w", font=("Arial", 14)).pack(side="left")
            entry = tk.Entry(row, font=("Arial", 14))
            entry.pack(side="left", fill="x", expand=True)
            entry.insert(0, self.config["SQL"].get(key, ""))
            self.entries[key] = entry
        tk.Label(frame, text="Note: Week end date is assumed to be Sunday.", font=("Arial", 12, "italic")).pack(pady=5)
        tk.Button(frame, text="Save Settings", command=self.save_settings, font=("Arial", 14), width=20).pack(pady=10)
        tk.Button(frame, text="Configure Signatures", command=self.open_signature_settings, font=("Arial", 14), width=25).pack(pady=10)

    def save_settings(self):
        for key, entry in self.entries.items():
            self.config["SQL"][key] = entry.get().strip()
        save_config(self.config)
        messagebox.showinfo("Settings", "SQL settings saved successfully.")

    def open_signature_settings(self):
        SignatureSettingsWindow(self, self.config)

# -----------------------
# Mapping Tab (for loadsheet cell mapping)
class MappingTab(tk.Frame):
    def __init__(self, parent, config):
        super().__init__(parent)
        self.config = config
        self.create_ui()

    def create_ui(self):
        frame = tk.Frame(self, padx=20, pady=20)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text="Loadsheet Mapping", font=("Arial", 16, "bold")).pack(pady=10)
        mapping_fields = [
            ("Date Cell", "date_cell"),
            ("Load Number Cell", "load_cell"),
            ("Collection Cell", "collection_cell"),
            ("Destination Cell", "destination_cell")
        ]
        self.mapping_entries = {}
        if "LOADSHEET_MAPPING" not in self.config:
            self.config["LOADSHEET_MAPPING"] = {"date_cell": "C6", "load_cell": "G6", "collection_cell": "B9", "destination_cell": "F9"}
        for label_text, key in mapping_fields:
            row = tk.Frame(frame)
            row.pack(fill="x", pady=5)
            tk.Label(row, text=f"{label_text}:", width=20, anchor="w", font=("Arial", 14)).pack(side="left")
            entry = tk.Entry(row, font=("Arial", 14))
            entry.pack(side="left", fill="x", expand=True)
            entry.insert(0, self.config["LOADSHEET_MAPPING"].get(key, ""))
            self.mapping_entries[key] = entry
        tk.Button(frame, text="Save Mapping", command=self.save_mapping, font=("Arial", 14), width=20).pack(pady=10)

    def save_mapping(self):
        for key, entry in self.mapping_entries.items():
            self.config["LOADSHEET_MAPPING"][key] = entry.get().strip()
        save_config(self.config)
        messagebox.showinfo("Mapping", "Loadsheet mapping saved successfully.")

# -----------------------
# Timesheets Tab (with working hours input)
class TimesheetsTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.create_ui()

    def create_ui(self):
        top_frame = tk.Frame(self)
        top_frame.pack(fill="x", padx=10, pady=10)
        tk.Label(top_frame, text="Select Week Ending Date (Sunday):", font=("Arial", 14)).pack(side="left", padx=5)
        default_date = upcoming_sunday().strftime("%Y-%m-%d")
        self.week_entry = DateEntry(top_frame, date_pattern="yyyy-mm-dd", font=("Arial", 14))
        self.week_entry.set_date(default_date)
        self.week_entry.pack(side="left", padx=5)
        tk.Button(top_frame, text="Generate Timesheet", command=self.generate_timesheet, font=("Arial", 14), width=20).pack(side="left", padx=5)
        tk.Button(top_frame, text="Convert Timesheets to PDF", command=self.convert_timesheets_to_pdf, font=("Arial", 14), width=25).pack(side="left", padx=5)
        # Working hours input for Monday-Sunday
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.hours = {}
        hours_frame = tk.Frame(self)
        hours_frame.pack(fill="x", padx=10, pady=10)
        for day in days:
            row = tk.Frame(hours_frame)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=f"{day} (Start - End):", font=("Arial", 12), width=15, anchor="w").pack(side="left")
            start_var = tk.StringVar()
            end_var = tk.StringVar()
            tk.Entry(row, textvariable=start_var, font=("Arial", 12), width=10).pack(side="left", padx=5)
            tk.Entry(row, textvariable=end_var, font=("Arial", 12), width=10).pack(side="left", padx=5)
            self.hours[day] = (start_var, end_var)

    def generate_timesheet(self):
        week_end_date = self.week_entry.get_date().strftime("%Y-%m-%d")
        prefix = f"Timesheet_{week_end_date}"
        excel_path = os.path.join(TIMESHEET_FOLDER, prefix + ".xlsx")
        if not os.path.exists(TEMPLATE_TIMESHEET):
            messagebox.showerror("Template Missing", f"Timesheet template not found at {TEMPLATE_TIMESHEET}")
            logger.error(f"Timesheet template missing: {TEMPLATE_TIMESHEET}")
            return
        try:
            wb = load_workbook(TEMPLATE_TIMESHEET)
            ws = wb.active
            ws["A1"] = f"Timesheet for week ending {week_end_date}"
            # Example: Map working hours to cells B4-B10 (adjust as needed)
            day_cells = {"Monday": "B4", "Tuesday": "B5", "Wednesday": "B6",
                         "Thursday": "B7", "Friday": "B8", "Saturday": "B9", "Sunday": "B10"}
            for day, cell in day_cells.items():
                start, end = self.hours[day]
                ws[cell] = f"{start.get()} - {end.get()}"
            wb.save(excel_path)
            messagebox.showinfo("Timesheet", f"Excel timesheet saved to:\n{excel_path}")
            logger.info(f"Saved timesheet: {excel_path}")
        except Exception as e:
            logger.error(f"Error generating timesheet: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to generate timesheet:\n{e}")

    def convert_timesheets_to_pdf(self):
        converted = 0
        for fname in os.listdir(TIMESHEET_FOLDER):
            if fname.endswith(".xlsx"):
                fpath = os.path.join(TIMESHEET_FOLDER, fname)
                pdf = convert_excel_to_pdf(fpath, TIMESHEET_FOLDER)
                if pdf:
                    converted += 1
        messagebox.showinfo("Timesheet PDF Conversion", f"Converted {converted} timesheet(s) to PDF.")

# -----------------------
# Loads Tab
class LoadsTab(tk.Frame):
    def __init__(self, parent, db_manager, config):
        super().__init__(parent)
        self.db_manager = db_manager
        self.config = config
        self.load_data = []
        self.create_ui()

    def create_ui(self):
        top_frame = tk.Frame(self)
        top_frame.pack(fill="x", padx=10, pady=10)
        tk.Label(top_frame, text="Select Week Ending Date (Sunday):", font=("Arial", 14)).pack(side="left", padx=5)
        default_date = upcoming_sunday().strftime("%Y-%m-%d")
        self.week_entry = DateEntry(top_frame, date_pattern="yyyy-mm-dd", font=("Arial", 14))
        self.week_entry.set_date(default_date)
        self.week_entry.pack(side="left", padx=5)
        tk.Button(top_frame, text="Load Week", command=self.load_week, font=("Arial", 14), width=15).pack(side="left", padx=5)
        self.add_sig_var = tk.IntVar()
        tk.Checkbutton(top_frame, text="Add Signatures", variable=self.add_sig_var, font=("Arial", 14)).pack(side="left", padx=5)
        tk.Button(top_frame, text="Generate Loadsheet for All", command=self.generate_loadsheet_all, font=("Arial", 14), width=25).pack(side="left", padx=5)
        tk.Button(top_frame, text="Convert Loadsheets to PDF", command=self.convert_loadsheets_to_pdf, font=("Arial", 14), width=25).pack(side="left", padx=5)
        columns = ("dwjdate", "collection", "destination", "car_count")
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        self.tree.heading("dwjdate", text="Date")
        self.tree.heading("collection", text="Collection")
        self.tree.heading("destination", text="Destination")
        self.tree.heading("car_count", text="Cars")
        for col in columns:
            self.tree.column(col, width=150)
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.bind("<Double-1>", lambda e: self.generate_loadsheet())
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Generate Loadsheet", command=self.generate_loadsheet, font=("Arial", 14), width=20).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Refresh Loads", command=self.load_week, font=("Arial", 14), width=20).pack(side="left", padx=5)

    def load_week(self):
        week_end_date = self.week_entry.get_date().strftime("%Y-%m-%d")
        start_date = (self.week_entry.get_date() - timedelta(days=6)).strftime("%Y%m%d")
        end_date = self.week_entry.get_date().strftime("%Y%m%d")
        self.load_data = self.db_manager.fetch_loads(start_date, end_date)
        self.tree.delete(*self.tree.get_children())
        for load in self.load_data:
            try:
                fdate = datetime.strptime(load["dwjdate"], "%Y%m%d").strftime("%A / %d/%m/%Y")
            except Exception:
                fdate = load["dwjdate"]
            collection = load.get("dwjname", "")
            dest = self.db_manager.fetch_destination(load["dwjload"])
            destination = dest.get("dwjname", "") if dest else ""
            car_count = load.get("dwjvehs", "0")
            self.tree.insert("", "end", iid=load["dwjload"], values=(fdate, collection, destination, car_count))
        self.winfo_toplevel().status_label.config(text=f"Loaded {len(self.load_data)} collection loads")

    def generate_loadsheet(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a load.")
            return
        load_id = selected[0]
        load = next((ld for ld in self.load_data if ld["dwjload"] == load_id), None)
        if not load:
            messagebox.showerror("Error", "Selected load not found in data.")
            return
        self._generate_loadsheet_for_load(load)

    def generate_loadsheet_all(self):
        if not self.load_data:
            messagebox.showwarning("No Loads", "No loads to generate loadsheets for.")
            return
        for load in self.load_data:
            self._generate_loadsheet_for_load(load)
        messagebox.showinfo("Loadsheet", "Loadsheet files generated for all loads.")
        self.load_week()

    def _generate_loadsheet_for_load(self, load):
        collection = load.get("dwjname", "").replace(" ", "_")
        prefix = f"{load['dwjdate']}_{load['dwjload']}_{collection}"
        excel_path = os.path.join(LOADSHEET_FOLDER, prefix + ".xlsx")
        if not os.path.exists(TEMPLATE_LOADSHEET):
            messagebox.showerror("Template Missing", f"Loadsheet template not found at {TEMPLATE_LOADSHEET}")
            logger.error(f"Loadsheet template missing: {TEMPLATE_LOADSHEET}")
            return
        try:
            wb = load_workbook(TEMPLATE_LOADSHEET)
            ws = wb.active
            mapping = self.config.get("LOADSHEET_MAPPING", {})
            date_cell = mapping.get("date_cell", "C6")
            load_cell = mapping.get("load_cell", "G6")
            coll_cell = mapping.get("collection_cell", "B9")
            dest_cell = mapping.get("destination_cell", "F9")
            ws[date_cell] = datetime.strptime(load["dwjdate"], "%Y%m%d").strftime("%A / %d/%m/%Y")
            ws[load_cell] = load["dwjload"]
            ws[coll_cell] = load.get("dwjname", "")
            dest = self.db_manager.fetch_destination(load["dwjload"])
            ws[dest_cell] = dest.get("dwjname", "") if dest else ""
            if self.add_sig_var.get():
                add_signatures(excel_path, load, self.config)
            wb.save(excel_path)
            logger.info(f"Saved loadsheet: {excel_path}")
        except Exception as e:
            logger.error(f"Error generating loadsheet for load {load['dwjload']}: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to generate loadsheet:\n{e}")

    def convert_loadsheets_to_pdf(self):
        converted = 0
        for fname in os.listdir(LOADSHEET_FOLDER):
            if fname.endswith(".xlsx"):
                fpath = os.path.join(LOADSHEET_FOLDER, fname)
                pdf = convert_excel_to_pdf(fpath, LOADSHEET_FOLDER)
                if pdf:
                    converted += 1
        messagebox.showinfo("Loadsheet PDF Conversion", f"Converted {converted} loadsheet(s) to PDF.")

# -----------------------
# Main Application Window
class MainWindow(tk.Tk):
    def __init__(self, db_manager, config):
        super().__init__()
        self.title("Load & Timesheet Manager")
        self.db_manager = db_manager
        self.config = config
        if self.config["Window"].get("geometry"):
            self.geometry(self.config["Window"]["geometry"])
        else:
            self.geometry("1200x800")
        self.initUI()

    def initUI(self):
        tab_control = ttk.Notebook(self)
        self.loads_tab = LoadsTab(tab_control, self.db_manager, self.config)
        self.timesheets_tab = TimesheetsTab(tab_control)
        self.mapping_tab = MappingTab(tab_control, self.config)
        self.settings_tab = SettingsTab(tab_control, self.config)
        tab_control.add(self.loads_tab, text="Loads")
        tab_control.add(self.timesheets_tab, text="Timesheet")
        tab_control.add(self.mapping_tab, text="Mapping")
        tab_control.add(self.settings_tab, text="Settings")
        tab_control.pack(expand=True, fill="both")
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Convert Loadsheets to PDF", command=self.loads_tab.convert_loadsheets_to_pdf, font=("Arial", 14), width=25).pack(side="left", padx=5)
        self.status_label = ttk.Label(self, text="Ready", font=("Arial", 14))
        self.status_label.pack(pady=10)

    def on_close(self):
        self.config["Window"]["geometry"] = self.geometry()
        save_config(self.config)
        if self.db_manager.conn:
            self.db_manager.conn.close()
        self.destroy()

# -----------------------
# Main Application Entry Point
def main():
    root = tk.Tk()
    root.withdraw()
    settings_dialog = SettingsDialog(root)
    root.wait_window(settings_dialog)
    db_manager = settings_dialog.get_db_manager()
    if not db_manager:
        sys.exit(1)
    config = load_config()
    root.deiconify()
    main_win = MainWindow(db_manager, config)
    main_win.protocol("WM_DELETE_WINDOW", main_win.on_close)
    main_win.mainloop()

if __name__ == "__main__":
    main()
